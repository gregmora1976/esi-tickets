from flask import Flask, render_template, jsonify, request, send_file, abort, redirect, url_for
from pathlib import Path
import json, webbrowser, os, urllib.request, urllib.parse
from datetime import datetime

APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / 'data'
CONFIG_FILE = DATA_DIR / 'config.json'
TICKETS_SUB = 'tickets'
FILES_SUB = 'fichiers'
SUPABASE_URL = (os.getenv("SUPABASE_URL") or "").rstrip("/")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY") or ""
SUPABASE_BUCKET = os.getenv("SUPABASE_BUCKET", "uploads")

app = Flask(__name__, template_folder='templates', static_folder='static')

def safe_filename(name):
    """Nettoie le nom du fichier pour Supabase tout en gardant le vrai nom affiché côté appli."""
    name = str(name or "fichier")
    return "".join(
        c if c.isalnum() or c in "._-" else "_"
        for c in name
    )


def supabase_upload_bytes(storage_path, content, content_type="application/octet-stream"):
    """Envoie un fichier dans Supabase Storage sans dépendre du SDK Python."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise RuntimeError("Variables SUPABASE_URL ou SUPABASE_SERVICE_KEY manquantes")

    safe_path = urllib.parse.quote(storage_path, safe="/")
    url = f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/{safe_path}"

    req = urllib.request.Request(
        url,
        data=content,
        method="POST",
        headers={
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "apikey": SUPABASE_KEY,
            "Content-Type": content_type or "application/octet-stream",
            "x-upsert": "true"
        }
    )

    print("[SUPABASE UPLOAD URL]", url)
    print("[SUPABASE UPLOAD BUCKET]", SUPABASE_BUCKET)
    print("[SUPABASE UPLOAD PATH]", safe_path)

    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return resp.read()
    except urllib.error.HTTPError as e:
        try:
            body = e.read().decode("utf-8", errors="replace")
        except Exception:
            body = ""
        print(f"[SUPABASE UPLOAD ERROR] HTTP {e.code} - {e.reason} - {body}")
        raise

def supabase_download_bytes(storage_path):
    """Télécharge un fichier depuis Supabase Storage."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise RuntimeError("Variables SUPABASE_URL ou SUPABASE_SERVICE_KEY manquantes")

    safe_path = urllib.parse.quote(storage_path, safe="/")
    url = f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/{safe_path}"

    req = urllib.request.Request(
        url,
        method="GET",
        headers={
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "apikey": SUPABASE_KEY
        }
    )

    with urllib.request.urlopen(req, timeout=60) as resp:
        return resp.read()


def supabase_signed_download_url(storage_path, expires_in=300):
    """Crée une URL signée Supabase Storage pour éviter de faire transiter le fichier par Render."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise RuntimeError("Variables SUPABASE_URL ou SUPABASE_SERVICE_KEY manquantes")

    safe_path = urllib.parse.quote(storage_path, safe="/")
    url = f"{SUPABASE_URL}/storage/v1/object/sign/{SUPABASE_BUCKET}/{safe_path}"

    payload = json.dumps({"expiresIn": int(expires_in)}).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=payload,
        method="POST",
        headers={
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "apikey": SUPABASE_KEY,
            "Content-Type": "application/json"
        }
    )

    with urllib.request.urlopen(req, timeout=30) as resp:
        body = json.loads(resp.read().decode("utf-8", errors="replace"))

    signed = body.get("signedURL") or body.get("signedUrl") or body.get("url")
    if not signed:
        raise RuntimeError(f"Réponse URL signée invalide : {body}")
    if signed.startswith("http"):
        return signed
    return SUPABASE_URL + "/storage/v1" + signed

def choose_shared_folder():
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        path = filedialog.askdirectory(title="Choisis le dossier partagé ESI Tickets")
        root.destroy()
        if path:
            return path
    except Exception:
        pass
    return ''

def load_config():
    if CONFIG_FILE.exists():
        try:
            return json.loads(CONFIG_FILE.read_text(encoding='utf-8'))
        except Exception:
            pass
    return {}

def save_config(cfg):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    CONFIG_FILE.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), encoding='utf-8')

def ensure_shared_root():
    root = APP_DIR
    (root / TICKETS_SUB).mkdir(parents=True, exist_ok=True)
    (root / FILES_SUB).mkdir(parents=True, exist_ok=True)
    return root

def tickets_dir():
    return ensure_shared_root() / TICKETS_SUB

def files_dir():
    return ensure_shared_root() / FILES_SUB

def ticket_file(ticket_id):
    return tickets_dir() / f'{ticket_id}.json'

def ticket_folder(ticket_id):
    path = files_dir() / ticket_id
    path.mkdir(parents=True, exist_ok=True)
    return path



def _as_text(value, default=''):
    if value is None:
        return default
    return str(value)


def supabase_rest_request(method, table, query='', payload=None, prefer=None):
    """Appelle l'API REST Supabase Database sans dépendre du SDK Python."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise RuntimeError("Variables SUPABASE_URL ou SUPABASE_SERVICE_KEY manquantes")

    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if query:
        url += "?" + query.lstrip('?')

    data = None
    headers = {
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "apikey": SUPABASE_KEY,
        "Content-Type": "application/json",
    }
    if prefer:
        headers["Prefer"] = prefer
    elif method.upper() in ("POST", "PATCH", "DELETE"):
        headers["Prefer"] = "return=representation"

    if payload is not None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")

    req = urllib.request.Request(url, data=data, method=method.upper(), headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            body = resp.read().decode("utf-8", errors="replace")
            if not body:
                return None
            return json.loads(body)
    except urllib.error.HTTPError as e:
        try:
            body = e.read().decode("utf-8", errors="replace")
        except Exception:
            body = ""
        print(f"[SUPABASE DB ERROR] {method} {url} -> HTTP {e.code} {e.reason} {body}")
        raise RuntimeError(f"Erreur Supabase DB HTTP {e.code}: {body or e.reason}")


def init_db():
    """Vérifie simplement que les tables Supabase répondent."""
    try:
        supabase_rest_request("GET", "tickets", "select=id&limit=1")
        print("[SUPABASE DB] Connexion OK")
    except Exception as e:
        print(f"[SUPABASE DB] Connexion impossible : {e}")


def _ticket_to_db_row(ticket):
    return {
        "id": _as_text(ticket.get("id")),
        "module": _as_text(ticket.get("module")),
        "status": _as_text(ticket.get("status")),
        "created_at": _as_text(ticket.get("createdAt")),
        "updated_at": _as_text(ticket.get("updatedAt")),
        "dossier": _as_text(ticket.get("dossier")),
        "ref": _as_text(ticket.get("ref")),
        "preteur": _as_text(ticket.get("preteur")),
        "expo": _as_text(ticket.get("expo")),
        "objet": _as_text(ticket.get("objet")),
        "charge_projet": _as_text(ticket.get("chargeProjet")),
        "type_caisse": _as_text(ticket.get("typeCaisse")),
        "dimensions": _as_text(ticket.get("dimensions")),
        "date_emballage": _as_text(ticket.get("dateEmballage")),
        "prix_devis": _as_text(ticket.get("prixDevis")),
        "date_rdv": _as_text(ticket.get("dateRdv")),
        "heure_rdv": _as_text(ticket.get("heureRdv")),
        "lieu_rdv": _as_text(ticket.get("lieuRdv")),
        "contact_rdv": _as_text(ticket.get("contactRdv")),
        "commentaire": _as_text(ticket.get("commentaire")),
        "validated_at": _as_text(ticket.get("validatedAt")),
        "raw_json": ticket,
    }


def _ticket_from_db_row(row):
    return {
        "id": row.get("id") or "",
        "module": row.get("module") or "",
        "status": row.get("status") or "",
        "createdAt": row.get("created_at") or "",
        "updatedAt": row.get("updated_at") or "",
        "dossier": row.get("dossier") or "",
        "ref": row.get("ref") or "",
        "preteur": row.get("preteur") or "-",
        "expo": row.get("expo") or "-",
        "objet": row.get("objet") or "-",
        "chargeProjet": row.get("charge_projet") or "-",
        "typeCaisse": row.get("type_caisse") or "-",
        "dimensions": row.get("dimensions") or "-",
        "dateEmballage": row.get("date_emballage") or "-",
        "prixDevis": row.get("prix_devis") or "-",
        "dateRdv": row.get("date_rdv") or "-",
        "heureRdv": row.get("heure_rdv") or "-",
        "lieuRdv": row.get("lieu_rdv") or "-",
        "contactRdv": row.get("contact_rdv") or "-",
        "commentaire": row.get("commentaire") or "",
        "validatedAt": row.get("validated_at") or "",
    }


def _fiche_to_db_row(ticket_id, fiche):
    return {
        "ticket_id": ticket_id,
        "longueur": _as_text(fiche.get("longueur")),
        "largeur": _as_text(fiche.get("largeur")),
        "hauteur": _as_text(fiche.get("hauteur")),
        "dimensions_ext": _as_text(fiche.get("dimensionsExt")),
        "prix_achat": _as_text(fiche.get("prixAchat")),
        "prix_cession": _as_text(fiche.get("prixCession")),
        "type_caisse_fiche": _as_text(fiche.get("typeCaisseFiche")),
        "bilan_carbone": _as_text(fiche.get("bilanCarbone")),
        "poids": _as_text(fiche.get("poids")),
        "choix_caissier": _as_text(fiche.get("choixCaissier")),
    }


def _fiche_from_db_row(row):
    return {
        "longueur": row.get("longueur") or "",
        "largeur": row.get("largeur") or "",
        "hauteur": row.get("hauteur") or "",
        "dimensionsExt": row.get("dimensions_ext") or "",
        "prixAchat": row.get("prix_achat") or "",
        "prixCession": row.get("prix_cession") or "",
        "typeCaisseFiche": row.get("type_caisse_fiche") or "",
        "bilanCarbone": row.get("bilan_carbone") or "",
        "poids": row.get("poids") or "",
        "choixCaissier": row.get("choix_caissier") or "",
    }


def _add_file_to_ticket(ticket, f):
    item = {
        "name": f.get("filename") or "",
        "size": f.get("size") or 0,
        "path": f.get("storage_path") or ""
    }
    if f.get("kind") == "gestionnaire":
        ticket.setdefault("managerSheets", []).append(item)
    else:
        ticket.setdefault("files", []).append(item)


def _attach_children(ticket):
    """Charge les enfants d'un seul ticket. Utilisé pour les actions ciblées."""
    tid = ticket.get("id")
    if not tid:
        return ticket

    safe_tid = urllib.parse.quote(tid, safe='')

    fiches = supabase_rest_request(
        "GET",
        "fiches",
        f"select=*&ticket_id=eq.{safe_tid}&limit=1"
    ) or []
    if fiches:
        ticket["fiche"] = _fiche_from_db_row(fiches[0])

    rows = supabase_rest_request(
        "GET",
        "ticket_files",
        f"select=*&ticket_id=eq.{safe_tid}&order=uploaded_at.asc"
    ) or []

    ticket["files"] = []
    ticket["managerSheets"] = []
    for f in rows:
        _add_file_to_ticket(ticket, f)

    return ticket


def _chunks(values, size=100):
    for i in range(0, len(values), size):
        yield values[i:i + size]


def _in_filter(values):
    # Format PostgREST : in.(DEM-001,DEM-002). Les ids internes ne contiennent pas de virgule.
    return urllib.parse.quote(",".join(values), safe=",-_")


def list_tickets(status=None, limit=None):
    query = "select=*&order=created_at.desc"
    if status:
        query += "&status=eq." + urllib.parse.quote(status, safe='')
    if limit:
        query += "&limit=" + str(int(limit))

    rows = supabase_rest_request("GET", "tickets", query) or []
    tickets = [_ticket_from_db_row(row) for row in rows]

    by_id = {t.get("id"): t for t in tickets if t.get("id")}
    ids = list(by_id.keys())
    if not ids:
        return tickets

    # Initialisation des listes pour éviter les champs absents côté interface
    for t in tickets:
        t["files"] = []
        t["managerSheets"] = []

    # Chargement groupé des fiches : au lieu de 1 requête par ticket
    for part in _chunks(ids):
        fiches = supabase_rest_request(
            "GET",
            "fiches",
            "select=*&ticket_id=in.(" + _in_filter(part) + ")"
        ) or []
        for f in fiches:
            tid = f.get("ticket_id")
            if tid in by_id:
                by_id[tid]["fiche"] = _fiche_from_db_row(f)

    # Chargement groupé des fichiers : au lieu de 1 requête par ticket
    for part in _chunks(ids):
        rows_files = supabase_rest_request(
            "GET",
            "ticket_files",
            "select=*&ticket_id=in.(" + _in_filter(part) + ")&order=uploaded_at.asc"
        ) or []
        for f in rows_files:
            tid = f.get("ticket_id")
            if tid in by_id:
                _add_file_to_ticket(by_id[tid], f)

    return tickets


def next_id(prefix):
    safe_prefix = urllib.parse.quote(prefix + '-*', safe='*-')
    rows = supabase_rest_request(
        "GET",
        "tickets",
        f"select=id&id=like.{safe_prefix}&order=id.desc&limit=5000"
    ) or []
    nums = []
    for row in rows:
        try:
            nums.append(int(str(row.get("id", "")).split('-')[1]))
        except Exception:
            pass
    mx = max(nums) if nums else 0
    return f"{prefix}-{mx+1:03d}"


def save_ticket(ticket):
    if not ticket.get("id"):
        raise RuntimeError("Ticket sans ID")

    ticket.setdefault("updatedAt", datetime.now().isoformat())

    # Upsert du ticket principal
    supabase_rest_request(
        "POST",
        "tickets",
        "on_conflict=id",
        [_ticket_to_db_row(ticket)],
        prefer="resolution=merge-duplicates,return=minimal"
    )

    ticket_id = ticket.get("id")
    safe_tid = urllib.parse.quote(ticket_id, safe='')

    # Fiche gestionnaire
    fiche = ticket.get("fiche") or {}
    if fiche:
        supabase_rest_request(
            "POST",
            "fiches",
            "on_conflict=ticket_id",
            [_fiche_to_db_row(ticket_id, fiche)],
            prefer="resolution=merge-duplicates,return=minimal"
        )
    else:
        supabase_rest_request("DELETE", "fiches", f"ticket_id=eq.{safe_tid}", prefer="return=minimal")

    # Fichiers : on remplace la liste associée au ticket
    supabase_rest_request("DELETE", "ticket_files", f"ticket_id=eq.{safe_tid}", prefer="return=minimal")

    file_rows = []
    for fs in ticket.get("files") or []:
        if fs and fs.get("name"):
            file_rows.append({
                "ticket_id": ticket_id,
                "kind": "demandeur",
                "filename": _as_text(fs.get("name")),
                "size": fs.get("size") or 0,
                "storage_path": _as_text(fs.get("path")),
            })

    manager_sheets = list(ticket.get("managerSheets") or [])
    legacy = ticket.get("managerSheet")
    if legacy and isinstance(legacy, dict) and legacy.get("name"):
        if not any(x.get("name") == legacy.get("name") for x in manager_sheets):
            manager_sheets.append(legacy)

    for fs in manager_sheets:
        if fs and fs.get("name"):
            file_rows.append({
                "ticket_id": ticket_id,
                "kind": "gestionnaire",
                "filename": _as_text(fs.get("name")),
                "size": fs.get("size") or 0,
                "storage_path": _as_text(fs.get("path")),
            })

    if file_rows:
        supabase_rest_request("POST", "ticket_files", "", file_rows, prefer="return=minimal")

    print("[SUPABASE DB] Ticket sauvegardé", ticket_id)


def load_ticket(ticket_id):
    safe_tid = urllib.parse.quote(ticket_id, safe='')
    rows = supabase_rest_request("GET", "tickets", f"select=*&id=eq.{safe_tid}&limit=1") or []
    if not rows:
        return None
    return _attach_children(_ticket_from_db_row(rows[0]))


# -----------------------------------------------------------------------------
# Référentiels métier : chargés de projet, clients, contacts
# -----------------------------------------------------------------------------
REFERENTIELS = {
    "project-managers": {
        "table": "project_managers",
        "allowed": ["nom", "email", "telephone", "actif"],
        "search": ["nom", "email", "telephone"],
        "required": ["nom"],
        "defaults": {"actif": True},
        "order": "nom.asc"
    },
    "clients": {
        "table": "clients",
        "allowed": ["nom", "adresse", "contact_nom", "contact_email", "contact_telephone", "actif"],
        "search": ["nom", "adresse", "contact_nom", "contact_email", "contact_telephone"],
        "required": ["nom"],
        "defaults": {"actif": True},
        "order": "nom.asc"
    },
    "contacts": {
        "table": "contacts",
        "allowed": ["nom", "email", "telephone", "client_nom", "fonction", "actif"],
        "search": ["nom", "email", "telephone", "client_nom", "fonction"],
        "required": ["nom"],
        "defaults": {"actif": True},
        "order": "nom.asc"
    }
}


def _referentiel_config(kind):
    cfg = REFERENTIELS.get(kind)
    if not cfg:
        abort(404)
    return cfg


def _clean_referentiel_payload(kind, data, partial=False):
    cfg = _referentiel_config(kind)
    data = data or {}
    payload = {}

    for field in cfg["allowed"]:
        if field in data:
            if field == "actif":
                payload[field] = bool(data.get(field))
            else:
                payload[field] = _as_text(data.get(field)).strip()

    if not partial:
        for field, value in cfg.get("defaults", {}).items():
            payload.setdefault(field, value)

        missing = [field for field in cfg.get("required", []) if not payload.get(field)]
        if missing:
            raise ValueError("Champ obligatoire manquant : " + ", ".join(missing))

    return payload


@app.route('/api/referentiels/<kind>', methods=['GET'])
def api_list_referentiel(kind):
    cfg = _referentiel_config(kind)
    q = (request.args.get('q') or '').strip()
    include_inactive = request.args.get('include_inactive') == '1'
    limit = request.args.get('limit') or '100'

    query = "select=*"
    if not include_inactive:
        query += "&actif=eq.true"
    if q:
        pattern = "*" + q.replace("*", "") + "*"
        parts = []
        for field in cfg["search"]:
            parts.append(f"{field}.ilike.{urllib.parse.quote(pattern, safe='*')}")
        query += "&or=(" + ",".join(parts) + ")"
    query += "&order=" + urllib.parse.quote(cfg.get("order", "nom.asc"), safe='.,')
    query += "&limit=" + urllib.parse.quote(str(limit), safe='')

    rows = supabase_rest_request("GET", cfg["table"], query) or []
    return jsonify(rows)


@app.route('/api/referentiels/<kind>', methods=['POST'])
def api_create_referentiel(kind):
    cfg = _referentiel_config(kind)
    data = request.get_json(silent=True) or {}
    try:
        payload = _clean_referentiel_payload(kind, data, partial=False)
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

    try:
        rows = supabase_rest_request("POST", cfg["table"], "", [payload], prefer="return=representation") or []
        return jsonify({'ok': True, 'item': rows[0] if rows else payload})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/referentiels/<kind>/<item_id>', methods=['PUT'])
def api_update_referentiel(kind, item_id):
    cfg = _referentiel_config(kind)
    data = request.get_json(silent=True) or {}
    try:
        payload = _clean_referentiel_payload(kind, data, partial=True)
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

    if not payload:
        return jsonify({'ok': False, 'error': 'Aucune donnée à modifier'}), 400

    safe_id = urllib.parse.quote(str(item_id), safe='')
    try:
        rows = supabase_rest_request("PATCH", cfg["table"], f"id=eq.{safe_id}", payload, prefer="return=representation") or []
        return jsonify({'ok': True, 'item': rows[0] if rows else payload})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/referentiels/<kind>/<item_id>/toggle', methods=['PATCH'])
def api_toggle_referentiel(kind, item_id):
    cfg = _referentiel_config(kind)
    data = request.get_json(silent=True) or {}
    actif = bool(data.get('actif'))
    safe_id = urllib.parse.quote(str(item_id), safe='')
    try:
        rows = supabase_rest_request("PATCH", cfg["table"], f"id=eq.{safe_id}", {"actif": actif}, prefer="return=representation") or []
        return jsonify({'ok': True, 'item': rows[0] if rows else {'id': item_id, 'actif': actif}})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/')
def index():
    return redirect(url_for('demandeur'))

@app.route('/demandeur')
def demandeur():
    return render_template('demandeur.html')

@app.route('/gestionnaire')
def gestionnaire():
    from flask import request, redirect, url_for
    if request.args.get('pwd') != '1234':
        return redirect(url_for('login'))
    return render_template('gestionnaire.html')

@app.route('/login', methods=['GET','POST'])
def login():
    from flask import request, redirect, render_template_string
    error = ''
    if request.method == 'POST':
        if request.form.get('password') == '1234':
            return redirect('/gestionnaire?pwd=1234')
        error = 'Mot de passe incorrect'
    return render_template_string("""<!DOCTYPE html>
<html lang='fr'>
<head>
<meta charset='UTF-8'>
<meta name='viewport' content='width=device-width, initial-scale=1.0'>
<title>Connexion gestionnaire</title>
<style>
body{font-family:Arial,Helvetica,sans-serif;background:linear-gradient(180deg,#eef6fb 0%,#f6f8fb 100%);margin:0;display:flex;align-items:center;justify-content:center;height:100vh;color:#1e293b}
.card{background:#fff;border:1px solid #dbe7f0;border-radius:20px;padding:28px;box-shadow:0 12px 30px rgba(15,23,42,.08);width:360px}
h1{margin:0 0 10px;font-size:28px} p{margin:0 0 18px;color:#64748b}
input{width:100%;padding:12px 14px;border:1px solid #dbe7f0;border-radius:14px;font-size:15px;box-sizing:border-box}
button{margin-top:14px;width:100%;padding:12px 14px;border:none;border-radius:14px;background:linear-gradient(135deg,#0ea5e9 0%, #0284c7 100%);color:#fff;font-weight:700;cursor:pointer}
.err{margin-top:12px;color:#b91c1c;font-size:13px}
</style>
</head>
<body>
  <form class='card' method='post'>
    <h1>Gestion Tickets</h1>
    <p>Accès protégé par mot de passe</p>
    <input type='password' name='password' placeholder='Mot de passe' autofocus />
    <button type='submit'>Entrer</button>
    {% if error %}<div class='err'>{{ error }}</div>{% endif %}
  </form>
</body>
</html>""", error=error)


@app.route('/reception')
def reception():
    return render_template('reception.html')

@app.route('/api/status')
def api_status():
    root = ensure_shared_root()
    return jsonify({'shared_path': str(root), 'mode': 'automatic_app_folder'})

@app.route('/api/tickets')
def api_tickets():
    status = request.args.get('status')
    limit = request.args.get('limit')
    tickets = list_tickets(status=status, limit=limit)
    return jsonify(tickets)

@app.route('/api/tickets', methods=['POST'])
def api_create_ticket():
    form = request.form
    module = form.get('module','')
    prefix = 'DEM' if module == 'Fiche de caisse' else ('DEV' if module == 'Demande de devis' else 'AV')
    ticket_id = next_id(prefix)
    ticket = {
        'id': ticket_id,
        'module': module,
        'status': 'Demande créée',
        'createdAt': datetime.now().isoformat(),
        'dossier': form.get('dossier',''),
        'ref': form.get('ref',''),
        'preteur': form.get('preteur','-') or '-',
        'expo': form.get('expo','-') or '-',
        'objet': form.get('objet','-') or '-',
        'chargeProjet': form.get('chargeProjet','-') or '-',
        'typeCaisse': form.get('typeCaisse','-') or '-',
        'dimensions': form.get('dimensions','-') or '-',
        'dateEmballage': form.get('dateEmballage','-') or '-',
        'prixDevis': form.get('prixDevis','-') or '-',
        'dateRdv': form.get('dateRdv','-') or '-',
        'heureRdv': form.get('heureRdv','-') or '-',
        'lieuRdv': form.get('lieuRdv','-') or '-',
        'contactRdv': form.get('contactRdv','-') or '-',
        'commentaire': form.get('commentaire',''),
        'files': [],
        'managerSheets': []
    }
    ticket_folder(ticket_id)  # conserve la création du dossier local historique

    for fs in request.files.getlist('files'):
        if not fs.filename:
            continue

        content = fs.read()
        clean_name = safe_filename(fs.filename)
        storage_path = f"{ticket_id}/{datetime.now().strftime('%Y%m%d%H%M%S')}_{clean_name}"

        try:
            supabase_upload_bytes(
                storage_path,
                content,
                fs.content_type
            )
        except Exception as e:
            print(f"[SUPABASE UPLOAD] Erreur : {e}")
            return jsonify({'ok': False, 'error': f'Erreur upload Supabase : {e}'}), 500

        ticket['files'].append({
            'name': fs.filename,
            'size': len(content),
            'path': storage_path
        })

    save_ticket(ticket)
    return jsonify({'ok': True, 'id': ticket_id})


@app.route('/api/tickets/<ticket_id>', methods=['PUT'])
def api_update_ticket(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404

    data = request.get_json(silent=True) or {}

    editable_fields = [
        'dossier',
        'ref',
        'preteur',
        'expo',
        'objet',
        'chargeProjet',
        'typeCaisse',
        'dimensions',
        'dateEmballage',
        'prixDevis',
        'dateRdv',
        'heureRdv',
        'lieuRdv',
        'contactRdv',
        'commentaire'
    ]

    for field in editable_fields:
        if field in data:
            ticket[field] = data.get(field, '')

    if 'expo' in data and 'objet' not in data:
        ticket['objet'] = data.get('expo', '')

    ticket['updatedAt'] = datetime.now().isoformat()
    save_ticket(ticket)
    return jsonify({'ok': True})

@app.route('/api/tickets/<ticket_id>/status', methods=['PATCH'])
def api_update_status(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404
    data = request.get_json(silent=True) or {}
    ticket['status'] = data.get('status', ticket.get('status'))
    ticket['updatedAt'] = datetime.now().isoformat()
    save_ticket(ticket)
    return jsonify({'ok': True})

@app.route('/api/tickets/<ticket_id>/manager-sheet', methods=['POST'])
def api_manager_sheet(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404

    files = request.files.getlist('files')
    if not files:
        single = request.files.get('file')
        if single:
            files = [single]

    valid_files = [fs for fs in files if fs and fs.filename]
    if not valid_files:
        return jsonify({'error': 'Fichier manquant'}), 400

    ticket_folder(ticket_id)  # conserve la création du dossier local historique
    manager_sheets = list(ticket.get('managerSheets') or [])
    legacy = ticket.get('managerSheet')
    if legacy and isinstance(legacy, dict) and legacy.get('name'):
        if not any(x.get('name') == legacy.get('name') for x in manager_sheets):
            manager_sheets.append(legacy)

    for fs in valid_files:
        content = fs.read()
        clean_name = safe_filename(fs.filename)
        storage_path = f"{ticket_id}/gestionnaire/{datetime.now().strftime('%Y%m%d%H%M%S')}_{clean_name}"

        try:
            supabase_upload_bytes(
                storage_path,
                content,
                fs.content_type
            )
        except Exception as e:
            print(f"[SUPABASE UPLOAD GESTIONNAIRE] Erreur : {e}")
            return jsonify({'ok': False, 'error': f'Erreur upload Supabase : {e}'}), 500

        manager_sheets = [x for x in manager_sheets if x.get('name') != fs.filename]
        manager_sheets.append({
            'name': fs.filename,
            'size': len(content),
            'path': storage_path
        })
    ticket['managerSheets'] = manager_sheets
    ticket['updatedAt'] = datetime.now().isoformat()
    save_ticket(ticket)
    return jsonify({'ok': True})

def _find_file_info(ticket, filename, kind):
    items = ticket.get('managerSheets') if kind == 'gestionnaire' else ticket.get('files')
    for f in items or []:
        if f.get('name') == filename:
            return f
    return None


def _redirect_to_signed_file(ticket_id, filename, kind):
    ticket = load_ticket(ticket_id)
    if not ticket:
        abort(404)

    file_info = _find_file_info(ticket, filename, kind)
    if not file_info:
        abort(404)

    storage_path = file_info.get('path')
    if not storage_path:
        abort(404)

    try:
        signed_url = supabase_signed_download_url(storage_path, expires_in=300)
    except Exception as e:
        # Secours : si l'URL signée échoue, on garde l'ancien comportement via Render.
        print(f"[SUPABASE SIGNED DOWNLOAD] Erreur, fallback Render : {e}")
        import io
        try:
            data = supabase_download_bytes(storage_path)
        except Exception as e2:
            print(f"[SUPABASE DOWNLOAD] Erreur : {e2}")
            abort(404)
        return send_file(io.BytesIO(data), as_attachment=True, download_name=filename)

    return redirect(signed_url)


@app.route('/api/tickets/<ticket_id>/download/<filename>')
def api_download_file(ticket_id, filename):
    return _redirect_to_signed_file(ticket_id, filename, 'demandeur')


@app.route('/api/tickets/<ticket_id>/download-sheet/<filename>')
def api_download_sheet(ticket_id, filename):
    return _redirect_to_signed_file(ticket_id, filename, 'gestionnaire')


@app.route('/api/tickets/<ticket_id>/fiche', methods=['GET'])
def api_get_fiche(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404
    return jsonify(ticket.get('fiche', {}))

@app.route('/api/tickets/<ticket_id>/fiche', methods=['POST'])
def api_save_fiche(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404
    data = request.get_json(silent=True) or {}
    longueur = data.get('longueur', '')
    largeur = data.get('largeur', '')
    hauteur = data.get('hauteur', '')
    dimensions_ext = " x ".join([v for v in [longueur, largeur, hauteur] if str(v).strip()])
    ticket['fiche'] = {
        'longueur': longueur,
        'largeur': largeur,
        'hauteur': hauteur,
        'dimensionsExt': dimensions_ext,
        'prixAchat': data.get('prixAchat', ''),
        'prixCession': data.get('prixCession', ''),
        'typeCaisseFiche': data.get('typeCaisseFiche', ''),
        'bilanCarbone': data.get('bilanCarbone', ''),
        'poids': data.get('poids', ''),
        'choixCaissier': data.get('choixCaissier', '')
    }
    save_ticket(ticket)
    return jsonify({'ok': True})



@app.route('/api/export/excel')
def api_export_excel():
    try:
        from openpyxl import Workbook
        import io
        import re
    except Exception:
        return jsonify({'error': "openpyxl non installé"}), 500

    tickets = list_tickets()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    ws.append([
        "ID","Module","Statut","Date création","Dossier / Client",
        "Réf / N° caisse","Chargé de projet","Projet / Expo",
        "Type de caisse","Dimensions","Prix devis",
        "Prix d'achat","Prix cession","Commentaire","Choix du caissier",
        "Date RDV","Heure RDV","Lieu RDV"
    ])

    def parse_euro(value):
        if value is None:
            return None
        txt = str(value).strip()
        if not txt or txt == '-':
            return None
        txt = txt.replace('\xa0', ' ').replace('€', '').replace(' ', '')
        txt = txt.replace(',', '.')
        txt = re.sub(r'[^0-9.\-]', '', txt)
        if not txt:
            return None
        try:
            return float(txt)
        except Exception:
            return None

    for t in tickets:
        fiche = t.get('fiche', {}) or {}
        ws.append([
            t.get('id',''),
            t.get('module',''),
            t.get('status',''),
            t.get('createdAt',''),
            t.get('dossier',''),
            t.get('ref',''),
            t.get('chargeProjet',''),
            t.get('expo') or t.get('objet',''),
            t.get('typeCaisse',''),
            t.get('dimensions',''),
            parse_euro(t.get('prixDevis','')),
            parse_euro(fiche.get('prixAchat','')),
            parse_euro(fiche.get('prixCession','')),
            t.get('commentaire',''),
            fiche.get('choixCaissier',''),
            t.get('dateRdv',''),
            t.get('heureRdv',''),
            t.get('lieuRdv','')
        ])

    for row in range(2, ws.max_row + 1):
        for col in [11, 12, 13]:
            ws.cell(row=row, column=col).number_format = '#,##0.00 €'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="tickets_esi.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



@app.route('/api/tickets/<ticket_id>/export-pdf')
def api_export_ticket_pdf(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404

    import io
    import textwrap

    def clean(value):
        if value is None or value == '':
            return '-'
        return str(value).replace('\r', ' ').replace('\n', ' ')

    def add_wrapped(lines, label, value):
        text = label + " : " + clean(value)
        for part in textwrap.wrap(text, width=82) or [text]:
            lines.append(part)

    lines = []
    lines.append("ESI TICKETS - DETAIL TICKET")
    lines.append("=" * 70)
    lines.append("")
    add_wrapped(lines, "ID", ticket.get('id'))
    add_wrapped(lines, "Module", ticket.get('module'))
    add_wrapped(lines, "Statut", ticket.get('status'))
    add_wrapped(lines, "Dossier / Client", ticket.get('dossier'))
    add_wrapped(lines, "Reference", ticket.get('ref'))
    add_wrapped(lines, "Charge de projet", ticket.get('chargeProjet'))
    add_wrapped(lines, "Projet / Expo", ticket.get('expo') or ticket.get('objet'))
    add_wrapped(lines, "Preteur", ticket.get('preteur'))
    add_wrapped(lines, "Type de caisse", ticket.get('typeCaisse'))
    add_wrapped(lines, "Dimensions", ticket.get('dimensions'))
    add_wrapped(lines, "Prix devis", ticket.get('prixDevis'))
    add_wrapped(lines, "Lieu RDV", ticket.get('lieuRdv'))
    add_wrapped(lines, "Date RDV", ticket.get('dateRdv'))
    add_wrapped(lines, "Heure RDV", ticket.get('heureRdv'))

    lines.append("")
    lines.append("COMMENTAIRE / INFORMATIONS")
    lines.append("-" * 70)
    commentaire = clean(ticket.get('commentaire'))
    for part in textwrap.wrap(commentaire, width=82) or ['-']:
        lines.append(part)

    fiche = ticket.get('fiche') or {}
    if fiche:
        lines.append("")
        lines.append("INFORMATIONS FICHE")
        lines.append("-" * 70)
        add_wrapped(lines, "Dimensions exterieures", fiche.get('dimensionsExt'))
        add_wrapped(lines, "Prix achat", fiche.get('prixAchat'))
        add_wrapped(lines, "Type caisse fiche", fiche.get('typeCaisseFiche'))
        add_wrapped(lines, "Bilan carbone", fiche.get('bilanCarbone'))
        add_wrapped(lines, "Poids", fiche.get('poids'))
        add_wrapped(lines, "Choix caissier", fiche.get('choixCaissier'))

    lines.append("")
    lines.append("DOCUMENTS DU DEMANDEUR")
    lines.append("-" * 70)
    files = ticket.get('files') or []
    if files:
        for f in files:
            lines.append("- " + clean(f.get('name')))
    else:
        lines.append("- Aucun document")

    manager_sheets = ticket.get('managerSheets') or []
    if manager_sheets:
        lines.append("")
        lines.append("DOCUMENTS GESTIONNAIRE")
        lines.append("-" * 70)
        for f in manager_sheets:
            lines.append("- " + clean(f.get('name')))

    lines.append("")
    lines.append("NOTES / ACTIONS A PREVOIR")
    lines.append("-" * 70)
    lines.append("")
    lines.append("_" * 70)
    lines.append("")
    lines.append("_" * 70)
    lines.append("")
    lines.append("_" * 70)

    def pdf_escape(value):
        value = str(value)
        value = value.replace("\\", "\\\\")
        value = value.replace("(", "\\(")
        value = value.replace(")", "\\)")
        return value

    page_width, page_height = 595, 842
    margin_left = 42
    y_start = 800
    line_height = 14
    max_lines = 53
    chunks = [lines[i:i+max_lines] for i in range(0, len(lines), max_lines)] or [["Ticket vide"]]

    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        None,
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>"
    ]

    page_refs = []
    for chunk in chunks:
        content_obj_num = len(objects) + 1
        content_lines = ["BT", "/F1 10 Tf", f"{margin_left} {y_start} Td"]
        first = True
        for line in chunk:
            if not first:
                content_lines.append(f"0 -{line_height} Td")
            first = False
            content_lines.append(f"({pdf_escape(line)}) Tj")
        content_lines.append("ET")

        stream = "\n".join(content_lines).encode("latin-1", errors="replace")
        objects.append(f"<< /Length {len(stream)} >>\nstream\n".encode("latin-1") + stream + b"\nendstream")

        page_obj_num = len(objects) + 1
        page = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_width} {page_height}] "
            f"/Resources << /Font << /F1 3 0 R >> >> /Contents {content_obj_num} 0 R >>"
        )
        objects.append(page.encode("latin-1"))
        page_refs.append(f"{page_obj_num} 0 R")

    objects[1] = f"<< /Type /Pages /Kids [{' '.join(page_refs)}] /Count {len(page_refs)} >>".encode("latin-1")

    pdf = io.BytesIO()
    pdf.write(b"%PDF-1.4\n")
    offsets = []

    for i, obj in enumerate(objects, start=1):
        offsets.append(pdf.tell())
        pdf.write(f"{i} 0 obj\n".encode("latin-1"))
        pdf.write(obj)
        pdf.write(b"\nendobj\n")

    xref_pos = pdf.tell()
    pdf.write(f"xref\n0 {len(objects)+1}\n".encode("latin-1"))
    pdf.write(b"0000000000 65535 f \n")
    for offset in offsets:
        pdf.write(f"{offset:010d} 00000 n \n".encode("latin-1"))

    trailer = f"trailer\n<< /Size {len(objects)+1} /Root 1 0 R >>\nstartxref\n{xref_pos}\n%%EOF"
    pdf.write(trailer.encode("latin-1"))
    pdf.seek(0)

    return send_file(
        pdf,
        as_attachment=True,
        download_name=f"{ticket.get('id','ticket')}.pdf",
        mimetype='application/pdf'
    )




@app.route('/api/restart')
def api_restart():
    import os
    os._exit(0)

@app.route('/splash')
def splash():
    return """
    <html>
    <head>
        <title>ESI Tickets</title>
        <style>
            body{margin:0;display:flex;justify-content:center;align-items:center;height:100vh;background:linear-gradient(180deg,#eef6fb,#f6f8fb);font-family:Arial;}
            .box{text-align:center;}
            img{width:120px;margin-bottom:20px;}
            h1{margin:0;color:#0284c7;}
            p{color:#64748b;}
        </style>
        <script>
            setTimeout(()=>{window.location.href="/demandeur";},1500);
        </script>
    </head>
    <body>
        <div class="box">
            <img id="splashLogo" src="/static/logo.jpg" onerror="
                const logos=['/static/logo.png','/static/logo%20esi.jpg'];
                const idx=Number(this.dataset.idx||0);
                if(idx<logos.length){this.dataset.idx=idx+1;this.src=logos[idx];}
                else{this.style.display='none';}
            ">
            <h1>ESI Tickets</h1>
            <p>Chargement en cours...</p>
        </div>
    </body>
    </html>
    """


@app.route('/api/tickets/<ticket_id>/validate-aller-voir', methods=['POST'])
def api_validate_aller_voir(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404
    ticket['status'] = 'En cours'
    ticket['validatedAt'] = datetime.now().isoformat()
    save_ticket(ticket)
    return jsonify({'ok': True})

@app.route('/api/tickets/<ticket_id>/calendar.ics')
def api_ticket_calendar_ics(ticket_id):
    ticket = load_ticket(ticket_id)
    if not ticket:
        return jsonify({'error': 'Ticket introuvable'}), 404

    date_rdv = ticket.get('dateRdv')
    heure_rdv = ticket.get('heureRdv')
    if not date_rdv or not heure_rdv or date_rdv == '-' or heure_rdv == '-':
        return jsonify({'error': 'Date/heure manquante'}), 400

    from datetime import timedelta
    start = datetime.fromisoformat(f"{date_rdv}T{heure_rdv}:00")
    end = start + timedelta(hours=2)

    def fmt(dt):
        return dt.strftime('%Y%m%dT%H%M%S')

    lieu = (ticket.get('lieuRdv') or '').strip()
    dossier = (ticket.get('dossier') or '').strip()
    summary = f"Aller voir - {lieu} - {dossier}"

    ics = f"""BEGIN:VCALENDAR
VERSION:2.0
BEGIN:VEVENT
DTSTART:{fmt(start)}
DTEND:{fmt(end)}
SUMMARY:{summary}
END:VEVENT
END:VCALENDAR"""

    from flask import Response
    return Response(ics, mimetype='text/calendar')



def open_browser():
    webbrowser.open('http://127.0.0.1:5050/splash')

ensure_shared_root()
init_db()

if __name__ == '__main__':
      app.run(host='127.0.0.1', port=5050, debug=False)
